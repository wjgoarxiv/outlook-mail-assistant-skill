from pathlib import Path
from zipfile import ZipFile

from openpyxl import load_workbook

from outlook_mail_assistant.report_exports import (
    export_task_candidates_to_csv,
    export_task_candidates_to_markdown,
    export_task_candidates_to_xlsx,
    export_markdown_to_docx,
)


def _sample_candidates():
    return [
        {
            "message_id": "a",
            "store_name": "Primary Mailbox",
            "folder_path": "Inbox",
            "subject": "Please review the spec",
            "sender_email": "owner@example.com",
            "received_at": "2026-04-01T10:00:00Z",
            "kind": "task",
            "confidence": "explicit",
            "reason": "request-like wording detected",
            "snippet": "Could you review the attached spec by Friday?",
            "needs_confirmation": False,
        },
        {
            "message_id": "b",
            "store_name": "Shared Mailbox",
            "folder_path": "Shared/Inbox",
            "subject": "Reminder: webinar on Friday",
            "sender_email": "shared@example.com",
            "received_at": "2026-04-01T12:00:00Z",
            "kind": "meeting",
            "confidence": "inferred",
            "reason": "meeting wording detected",
            "snippet": "Webinar starts Friday at 10:00.",
            "needs_confirmation": True,
        },
    ]


def test_export_task_candidates_to_markdown_writes_grouped_sections(tmp_path: Path):
    path = export_task_candidates_to_markdown(_sample_candidates(), tmp_path / "tasks.md")
    content = path.read_text(encoding="utf-8")

    assert "Executive Summary" in content
    assert "Immediate Actions" in content
    assert "Needs Confirmation" in content
    assert "Please review the spec" in content


def test_export_task_candidates_to_csv_writes_flat_table(tmp_path: Path):
    path = export_task_candidates_to_csv(_sample_candidates(), tmp_path / "tasks.csv")
    content = path.read_text(encoding="utf-8")

    assert "message_id,store_name,folder_path" in content
    assert "Please review the spec" in content


def test_export_task_candidates_to_xlsx_writes_workbook(tmp_path: Path):
    path = export_task_candidates_to_xlsx(_sample_candidates(), tmp_path / "tasks.xlsx")
    workbook = load_workbook(path)
    assert workbook.sheetnames[:4] == ["Summary", "Review Queue", "Needs Confirmation", "Immediate Actions"]
    summary = workbook["Summary"]
    review = workbook["Review Queue"]

    assert summary["A1"].value == "Outlook Mail Assistant Task Review"
    assert review["A1"].value == "Message ID"
    subjects = {review["D2"].value, review["D3"].value}
    assert "Please review the spec" in subjects
    assert "Reminder: webinar on Friday" in subjects


def test_export_markdown_to_docx_creates_repo_local_docx(tmp_path: Path):
    md_path = tmp_path / "tasks.md"
    md_path.write_text("# Test", encoding="utf-8")

    docx_path = export_markdown_to_docx(md_path, tmp_path / "tasks.docx")

    assert docx_path.exists()
    with ZipFile(docx_path) as archive:
        assert "word/document.xml" in archive.namelist()
