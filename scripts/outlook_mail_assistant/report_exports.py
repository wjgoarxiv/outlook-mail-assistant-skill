from __future__ import annotations

import csv
from datetime import datetime, timezone
from pathlib import Path
import shutil
import subprocess
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .docx_export import convert_markdown_to_docx
from .skill_paths import find_skill_script


HEADERS = [
    "message_id",
    "store_name",
    "folder_path",
    "subject",
    "sender_email",
    "received_at",
    "kind",
    "confidence",
    "reason",
    "snippet",
    "needs_confirmation",
]

DISPLAY_HEADERS = [
    "Message ID",
    "Mailbox",
    "Folder",
    "Subject",
    "Sender",
    "Received",
    "Action Type",
    "Confidence",
    "Detection Reason",
    "Snippet",
    "Needs Confirmation",
]

TITLE_FILL = PatternFill("solid", fgColor="203864")
HEADER_FILL = PatternFill("solid", fgColor="D9E2F3")
EXPLICIT_FILL = PatternFill("solid", fgColor="E2F0D9")
CONFIRM_FILL = PatternFill("solid", fgColor="FFF2CC")
THIN_BORDER = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)


def _sorted_candidates(candidates):
    def sort_key(item):
        return (
            item.get("needs_confirmation", False),
            item.get("received_at") or "",
            item.get("subject") or "",
        )

    return sorted(candidates, key=sort_key, reverse=True)


def _count_by(items, key):
    counts = {}
    for item in items:
        counts[item[key]] = counts.get(item[key], 0) + 1
    return counts


def export_task_candidates_to_markdown(candidates, output_path: Path) -> Path:
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    candidates = _sorted_candidates(candidates)
    explicit = [item for item in candidates if not item.get("needs_confirmation")]
    inferred = [item for item in candidates if item.get("needs_confirmation")]
    kind_counts = _count_by(candidates, "kind")

    lines = [
        "# Outlook Mail Assistant Task Review Report",
        "",
        f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}",
        "",
        "## Executive Summary",
        f"- Total candidates: {len(candidates)}",
        f"- Immediate actions: {len(explicit)}",
        f"- Needs confirmation: {len(inferred)}",
    ]
    for kind, count in sorted(kind_counts.items()):
        lines.append(f"- {kind}: {count}")

    lines.extend(
        [
            "",
            "## Immediate Actions",
            "",
            "| Subject | Sender | Received | Kind | Folder | Reason |",
            "|---|---|---|---|---|---|",
        ]
    )
    for item in explicit[:50]:
        lines.append(
            f"| {item['subject']} | {item['sender_email']} | {item['received_at']} | {item['kind']} | {item['folder_path']} | {item['reason']} |"
        )

    lines.extend(
        [
            "",
            "## Needs Confirmation",
            "",
            "| Subject | Sender | Received | Kind | Folder | Reason |",
            "|---|---|---|---|---|---|",
        ]
    )
    for item in inferred[:50]:
        lines.append(
            f"| {item['subject']} | {item['sender_email']} | {item['received_at']} | {item['kind']} | {item['folder_path']} | {item['reason']} |"
        )

    lines.extend(
        [
            "",
            "## Notes",
            "- The workbook/CSV versions contain the full review queue.",
            "- This DOCX/Markdown report is intentionally condensed for human review.",
        ]
    )
    output_path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    return output_path


def export_task_candidates_to_csv(candidates, output_path: Path) -> Path:
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=HEADERS)
        writer.writeheader()
        writer.writerows(_sorted_candidates(candidates))
    return output_path


def _apply_header_style(sheet, row_idx: int):
    for cell in sheet[row_idx]:
        cell.font = Font(bold=True, color="1F1F1F")
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def _apply_data_style(sheet, start_row: int, end_row: int):
    for row in sheet.iter_rows(min_row=start_row, max_row=end_row):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = THIN_BORDER


def _fit_columns(sheet):
    widths = {
        "A": 16,
        "B": 22,
        "C": 24,
        "D": 42,
        "E": 28,
        "F": 22,
        "G": 16,
        "H": 14,
        "I": 28,
        "J": 60,
        "K": 18,
    }
    for key, value in widths.items():
        sheet.column_dimensions[key].width = value


def _write_detail_sheet(sheet, title: str, candidates):
    sheet.title = title
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:K{max(len(candidates) + 1, 2)}"
    sheet.append(DISPLAY_HEADERS)
    _apply_header_style(sheet, 1)
    for item in candidates:
        row = [item.get(header) for header in HEADERS]
        row[-1] = "Yes" if item.get("needs_confirmation") else "No"
        sheet.append(row)
    _apply_data_style(sheet, 2, max(len(candidates) + 1, 2))
    _fit_columns(sheet)
    for row in range(2, len(candidates) + 2):
        if sheet[f"H{row}"].value == "explicit":
            for col in range(1, 12):
                sheet.cell(row=row, column=col).fill = EXPLICIT_FILL
        if sheet[f"K{row}"].value == "Yes":
            for col in range(1, 12):
                sheet.cell(row=row, column=col).fill = CONFIRM_FILL


def export_task_candidates_to_xlsx(candidates, output_path: Path) -> Path:
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    candidates = _sorted_candidates(candidates)
    explicit = [item for item in candidates if not item.get("needs_confirmation")]
    inferred = [item for item in candidates if item.get("needs_confirmation")]

    workbook = Workbook()
    summary = workbook.active
    if summary is None:
        raise RuntimeError("Failed to create XLSX summary sheet")
    summary.title = "Summary"
    summary["A1"] = "Outlook Mail Assistant Task Review"
    summary["A1"].font = Font(bold=True, size=16, color="FFFFFF")
    summary["A1"].fill = TITLE_FILL
    summary.merge_cells("A1:D1")
    summary["A3"] = "Generated"
    summary["B3"] = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    summary["A5"] = "Metric"
    summary["B5"] = "Value"
    _apply_header_style(summary, 5)
    summary["A6"] = "Total candidates"
    summary["B6"] = "=COUNTA('Review Queue'!A2:A1048576)"
    summary["A7"] = "Immediate actions"
    summary["B7"] = '=COUNTIF(\'Review Queue\'!K:K,"No")'
    summary["A8"] = "Needs confirmation"
    summary["B8"] = '=COUNTIF(\'Review Queue\'!K:K,"Yes")'
    summary["A10"] = "Task"
    summary["B10"] = '=COUNTIF(\'Review Queue\'!G:G,"task")'
    summary["A11"] = "Deadline"
    summary["B11"] = '=COUNTIF(\'Review Queue\'!G:G,"deadline")'
    summary["A12"] = "Meeting"
    summary["B12"] = '=COUNTIF(\'Review Queue\'!G:G,"meeting")'
    summary["A13"] = "Decision signal"
    summary["B13"] = '=COUNTIF(\'Review Queue\'!G:G,"decision_signal")'
    summary.column_dimensions["A"].width = 24
    summary.column_dimensions["B"].width = 18

    confirm_sheet = workbook.create_sheet("Needs Confirmation")
    _write_detail_sheet(confirm_sheet, "Needs Confirmation", inferred)
    explicit_sheet = workbook.create_sheet("Immediate Actions")
    _write_detail_sheet(explicit_sheet, "Immediate Actions", explicit)
    review = workbook.create_sheet("Review Queue", 1)
    _write_detail_sheet(review, "Review Queue", candidates)

    resolved = output_path.resolve()
    workbook.save(str(resolved))
    workbook.close()

    recalc_script = find_skill_script(
        ("007_xlsx", "recalc.py"),
        ("07_xlsx", "recalc.py"),
        ("document-skills", "xlsx", "recalc.py"),
    )
    if recalc_script is not None and shutil.which("soffice"):
        subprocess.run([sys.executable, str(recalc_script), str(resolved), "30"], capture_output=True, text=True, check=False)
    return resolved


def export_markdown_to_docx(markdown_path: Path, output_path: Path) -> Path:
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    return convert_markdown_to_docx(markdown_path, output_path)
